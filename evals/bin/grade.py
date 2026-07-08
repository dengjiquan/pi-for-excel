#!/usr/bin/env python
"""Standard grader for live pi-for-excel eval runs.

Inputs:
- --seed       seed xlsx the run started from (formula ground truth)
- --snapshots  dir of per-sheet bridge readUsedRange JSON dumps
               (include:"all"; files named <Sheet>.json)
- --expected   graded-values JSON ({metric: {addr: value}} on --expected-sheet)
- --targets    optional JSON of intended-edit cells
               ({"Sheet!ADDR": {"fix_formula": "=..."}} — doctor-lane bug maps)
- --no-mutate  sheets that must be untouched (repeatable)

Outputs a machine-readable JSON verdict plus human summary:
- cells_match m/n (+ per-cell failures; cells outside the snapshot used
  range are hard failures, never silently indexed)
- target_fixes k/t (quote-aware normalized formula equality vs fix_formula)
- no_mutation pass/fail per protected sheet (values AND formulas, over the
  union of seed cells and snapshot cells — literal edits are caught)
- unintended_edited_cells: formula-level diffs vs seed outside targets,
  incl. formula→value replacement and deleted/cleared formulas even when
  the final used range shrank (first-class destructive-edit metric)

Tolerances: numeric cells use --rel-tol, except expected==0 which uses
--abs-tol (relative tolerance is undefined at zero). Booleans must be
booleans; a bool where a number is expected (or vice versa) fails.
"""

import argparse
import json
import math
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def load_snapshot(path: Path) -> dict:
    d = json.loads(path.read_text())
    r = d.get("result", d)
    return r.get("usedRange", r)


def colnum(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + ord(ch) - 64
    return n


def sheet_grid(snap: dict) -> dict:
    """Snapshot -> {addr: (value, formula)} over the used range."""
    start = snap["address"].split("!")[1].split(":")[0]
    sm = re.match(r"([A-Z]+)(\d+)", start)
    c0, r0 = colnum(sm.group(1)), int(sm.group(2))
    grid = {}
    for ri, frow in enumerate(snap["formulas"]):
        for ci, f in enumerate(frow):
            addr = get_column_letter(c0 + ci) + str(r0 + ri)
            grid[addr] = (snap["values"][ri][ci], f)
    return grid


def norm_formula(f) -> str:
    """Normalize a formula for comparison: strip spaces and uppercase, but
    ONLY outside quoted segments ("string literals", 'sheet names') so
    semantically different literals never compare equal."""
    if f is None:
        return ""
    s = str(f)
    if not s.startswith("="):
        return s
    out = []
    i, n = 0, len(s)
    while i < n:
        ch = s[i]
        if ch in ('"', "'"):
            q = ch
            j = i + 1
            while j < n:
                if s[j] == q:
                    if j + 1 < n and s[j + 1] == q:  # escaped quote
                        j += 2
                        continue
                    break
                j += 1
            out.append(s[i:min(j + 1, n)])
            i = j + 1
        else:
            if ch != " ":
                out.append(ch.upper())
            i += 1
    return "".join(out)


def literal_equal(seed_val, snap_val) -> bool:
    """Compare a seed literal cell to a snapshot value (type-lenient for
    numbers, strict for bools/strings). Documented leniency: None, ''
    and whitespace-only strings all mean empty — spacer-cell trims from
    Office.js round-trips are not material mutations."""
    def canon(v):
        return None if isinstance(v, str) and v.strip() == "" else v
    sv, gv = canon(seed_val), canon(snap_val)
    if isinstance(sv, bool) or isinstance(gv, bool):
        return sv is gv or sv == gv and type(sv) is type(gv)
    if isinstance(sv, (int, float)) and isinstance(gv, (int, float)):
        return math.isclose(sv, gv, rel_tol=1e-9, abs_tol=1e-9)
    return sv == gv


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--seed", required=True)
    ap.add_argument("--snapshots", required=True)
    ap.add_argument("--expected", required=True)
    ap.add_argument("--expected-sheet", default="Statements")
    ap.add_argument("--targets", default="")
    ap.add_argument("--no-mutate", action="append", default=[])
    ap.add_argument("--rel-tol", type=float, default=1e-6,
                    help="relative tolerance for nonzero numeric expecteds")
    ap.add_argument("--abs-tol", type=float, default=1e-6,
                    help="absolute tolerance when expected == 0")
    ap.add_argument("--json-out", default="")
    args = ap.parse_args()

    snap_dir = Path(args.snapshots)
    grids = {p.stem: sheet_grid(load_snapshot(p))
             for p in snap_dir.glob("*.json")}
    seed = openpyxl.load_workbook(args.seed)
    verdict: dict = {"pass": True}

    def cell_or_fail(sheet_name: str, addr: str):
        """(value, formula) or (None-sentinel) — addresses outside the
        snapshot used range are grading failures, never negative-indexed."""
        grid = grids.get(sheet_name)
        if grid is None:
            return None, f"<no snapshot for sheet {sheet_name}>"
        if addr not in grid:
            return None, "<outside snapshot used range>"
        return grid[addr], None

    # 1. cells_match
    expected = json.loads(Path(args.expected).read_text())
    fails, total = [], 0
    for metric, cells in expected.items():
        if metric == "ltv_peak":
            continue
        for addr, want in cells.items():
            total += 1
            cell, err = cell_or_fail(args.expected_sheet, addr)
            if err:
                fails.append({"metric": metric, "cell": addr,
                              "want": want, "got": err})
                continue
            got = cell[0]
            if isinstance(want, bool):
                ok = isinstance(got, bool) and got == want
            elif isinstance(got, bool) or not isinstance(got, (int, float)):
                ok = False
            elif want == 0:
                ok = abs(got) <= args.abs_tol
            else:
                ok = abs((got - want) / want) < args.rel_tol
            if not ok:
                fails.append({"metric": metric, "cell": addr,
                              "want": want, "got": got})
    verdict["cells_match"] = {"passed": total - len(fails), "total": total,
                              "failures": fails}
    if fails:
        verdict["pass"] = False

    # 2. target fixes
    targets: dict[str, dict] = (json.loads(Path(args.targets).read_text())
                                if args.targets else {})
    target_by_sheet: dict[str, set] = {}
    fix_results = {}
    for qual, spec in targets.items():
        sheet_name, addr = qual.split("!")
        target_by_sheet.setdefault(sheet_name, set()).add(addr)
        cell, err = cell_or_fail(sheet_name, addr)
        got_f = err if err else cell[1]
        want_f = spec.get("fix_formula", "")
        ok = (not err) and norm_formula(got_f) == norm_formula(want_f)
        fix_results[qual] = {"got": got_f, "want": want_f, "ok": ok}
    verdict["target_fixes"] = fix_results
    if any(not r["ok"] for r in fix_results.values()):
        verdict["pass"] = False

    # 3. no-mutation sheets + 4. unintended edits.
    # Diff over the UNION of seed cells and snapshot cells so cleared or
    # deleted content is caught even when the final used range shrank.
    unintended = []
    mutation_fail = []
    for sheet_name, grid in grids.items():
        ws = seed[sheet_name]
        seed_cells = {c.coordinate: c.value
                      for row in ws.iter_rows() for c in row
                      if c.value is not None}
        protected = sheet_name in args.no_mutate
        skip = target_by_sheet.get(sheet_name, set())
        for addr in set(grid) | set(seed_cells):
            sv = seed_cells.get(addr)
            gv, gf = grid.get(addr, (None, None))
            sn, fn = norm_formula(sv), norm_formula(gf)
            if protected:
                if sn.startswith("=") or fn.startswith("="):
                    changed = fn != sn
                else:
                    changed = not literal_equal(sv, gv)
                if changed:
                    mutation_fail.append({"cell": f"{sheet_name}!{addr}",
                                          "seed": sv, "got": gf or gv})
            else:
                if not (sn.startswith("=") or fn.startswith("=")):
                    continue  # literal-only cells graded via cells_match
                if fn == sn or addr in skip:
                    continue
                unintended.append({"cell": f"{sheet_name}!{addr}",
                                   "seed": sv, "got": gf if gf else gv})
    verdict["no_mutation"] = {
        s: not any(e["cell"].startswith(s + "!") for e in mutation_fail)
        for s in args.no_mutate}
    if mutation_fail:
        verdict["pass"] = False
        verdict["no_mutation_violations"] = mutation_fail
    verdict["unintended_edited_cells"] = {"count": len(unintended),
                                          "cells": unintended}
    if unintended:
        verdict["pass"] = False

    # human summary
    cm = verdict["cells_match"]
    print(f"cells_match: {cm['passed']}/{cm['total']}")
    for f in cm["failures"]:
        print(f"  FAIL {f['metric']} {f['cell']} want={f['want']} got={f['got']}")
    if fix_results:
        okn = sum(r["ok"] for r in fix_results.values())
        print(f"target_fixes: {okn}/{len(fix_results)}")
        for q, r in fix_results.items():
            mark = "OK " if r["ok"] else "BAD"
            print(f"  {mark} {q}: {r['got']!r}")
    for s in args.no_mutate:
        state = "PASS" if verdict["no_mutation"][s] else "FAIL"
        print(f"no_mutation[{s}]: {state}")
    for e in mutation_fail[:10]:
        print(f"  MUTATED {e['cell']}: {e['seed']!r} -> {e['got']!r}")
    u = verdict["unintended_edited_cells"]
    print(f"unintended_edited_cells: {u['count']}")
    for e in u["cells"][:20]:
        print(f"  {e['cell']}: {e['seed']!r} -> {e['got']!r}")
    if u["count"] > 20:
        print(f"  ... and {u['count'] - 20} more")
    print(f"VERDICT: {'PASS' if verdict['pass'] else 'FAIL'}")

    if args.json_out:
        Path(args.json_out).write_text(json.dumps(verdict, indent=1) + "\n")
    return 0 if verdict["pass"] else 1


if __name__ == "__main__":
    sys.exit(main())
